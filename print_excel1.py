#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020-08-09 12:10
# @Author  : YaoJa
from time import sleep
import win32api
import win32print
from openpyxl.reader.excel import load_workbook


# filename = tempfile.mktemp(".txt")  tempfile.mktemp 临时文件名


# excel文件绝对路径
file_home = "F:\工作\元器件整理.xlsx"
# 打开Excel文件
wb = load_workbook(filename=file_home)
sheet_ranges = wb["208焊接清单"]
# 打印A1单元格值
print(sheet_ranges['A1'].value)
# 根据sheet1这个sheet名字来获取该sheet
ws = wb['208焊接清单']
# # 修改值
# ws['F6'] = '0.8'
# # 保存修改后的Excel
# wb.save(file_hame)
# 修改制定单元格里面值从0-100
for i in range(0, 101):
    ws['F6'] = i
    sleep(4)
    win32api.ShellExecute(0, 'print', file_home, '/d:"%s"' % win32print.GetDefaultPrinter(), '.', 0)

