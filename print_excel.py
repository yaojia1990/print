#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020-08-09 11:54
# @Author  : YaoJa
import os
import os.path
from openpyxl.reader.excel import load_workbook


# excel文件绝对路径
file_hame = "F:\工作\元器件整理.xlsx"
# 打开Excel文件
wb = load_workbook(filename=file_hame)
sheet_ranges = wb["208焊接清单"]
# 打印A1单元格值
print(sheet_ranges['A1'].value)
# 根据sheet1这个sheet名字来获取该sheet
ws = wb['208焊接清单']
# 修改值
ws['F6'] = '0.8'
# 保存修改后的Excel
wb.save(file_hame)
